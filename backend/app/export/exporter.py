"""
Generic, report-agnostic exporters.

Any report (ledger, trial balance, P&L, receivables, ...) is just a
title + a list of columns + a list of row dicts, so one exporter
here serves all of them instead of writing a new one per report.
"""

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer
)
from reportlab.lib.styles import getSampleStyleSheet


def _row_values(columns, rows):
    values = []

    for row in rows:
        line = []

        for column in columns:
            renderer = column.get("format")
            raw = row.get(column["key"])
            line.append(renderer(raw) if renderer else raw)

        values.append(line)

    return values


def build_excel(
    title: str,
    columns: list[dict],
    rows: list[dict],
    company_name: str | None = None,
    footer: dict | None = None
) -> bytes:
    """
    columns: [{"key": "name", "label": "Ledger"}, ...]
    footer: optional {"label": "Total", "key": "amount", "value": 12345}
            style totals row shown under the table.
    """

    workbook = Workbook()
    sheet = workbook.active

    invalid_chars = r"\/?*[]:"
    safe_title = "".join(
        char for char in title if char not in invalid_chars
    ).strip()
    sheet.title = safe_title[:31] or "Report"

    header_fill = PatternFill(
        start_color="1F2937",
        end_color="1F2937",
        fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(*(Side(style="thin", color="CCCCCC"),) * 4)

    row_cursor = 1

    sheet.cell(
        row=row_cursor, column=1, value=company_name or title
    ).font = Font(bold=True, size=14)
    row_cursor += 1

    sheet.cell(row=row_cursor, column=1, value=title).font = Font(
        bold=True, size=11, color="555555"
    )
    row_cursor += 1

    sheet.cell(
        row=row_cursor,
        column=1,
        value=f"Generated on {datetime.now().strftime('%d-%b-%Y %H:%M')}"
    ).font = Font(size=9, italic=True, color="888888")
    row_cursor += 2

    header_row = row_cursor

    for col_index, column in enumerate(columns, start=1):
        cell = sheet.cell(
            row=header_row, column=col_index, value=column["label"]
        )
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_offset, values in enumerate(
        _row_values(columns, rows), start=1
    ):
        for col_index, value in enumerate(values, start=1):
            cell = sheet.cell(
                row=header_row + row_offset,
                column=col_index,
                value=value
            )
            cell.border = thin_border

            if isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "#,##0.00"

    last_row = header_row + len(rows)

    if footer:
        last_row += 1
        label_col = 1
        value_col = next(
            (
                index
                for index, column in enumerate(columns, start=1)
                if column["key"] == footer["key"]
            ),
            len(columns)
        )

        sheet.cell(row=last_row, column=label_col, value=footer["label"]).font = (
            Font(bold=True)
        )
        value_cell = sheet.cell(
            row=last_row, column=value_col, value=footer["value"]
        )
        value_cell.font = Font(bold=True)
        value_cell.number_format = "#,##0.00"

    for col_index, column in enumerate(columns, start=1):
        longest_value = len(column["label"])

        for row in rows:
            cell_value = row.get(column["key"])
            if cell_value is not None:
                longest_value = max(longest_value, len(str(cell_value)))

        width = min(max(12, longest_value + 4), 45)
        sheet.column_dimensions[get_column_letter(col_index)].width = width

    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_pdf(
    title: str,
    columns: list[dict],
    rows: list[dict],
    company_name: str | None = None,
    footer: dict | None = None
) -> bytes:
    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4) if len(columns) > 5 else A4,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
    )

    styles = getSampleStyleSheet()
    elements = [
        Paragraph(company_name or title, styles["Title"]),
        Paragraph(title, styles["Heading3"]),
        Paragraph(
            f"Generated on {datetime.now().strftime('%d-%b-%Y %H:%M')}",
            styles["Normal"]
        ),
        Spacer(1, 10),
    ]

    header = [column["label"] for column in columns]
    body = _row_values(columns, rows)

    for line in body:
        for index, value in enumerate(line):
            if isinstance(value, float):
                line[index] = f"{value:,.2f}"
            elif value is None:
                line[index] = ""

    table_data = [header] + body

    if footer:
        footer_row = ["" for _ in columns]
        footer_row[0] = footer["label"]
        value_col = next(
            (
                index
                for index, column in enumerate(columns)
                if column["key"] == footer["key"]
            ),
            len(columns) - 1
        )
        value = footer["value"]
        footer_row[value_col] = (
            f"{value:,.2f}" if isinstance(value, (int, float)) else value
        )
        table_data.append(footer_row)

    table = Table(table_data, repeatRows=1)

    style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2 if footer else -1),
         [colors.white, colors.HexColor("#F7F7F7")]),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]

    if footer:
        style.append(("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"))
        style.append(
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#EAEAEA"))
        )

    table.setStyle(TableStyle(style))
    elements.append(table)

    doc.build(elements)
    return buffer.getvalue()